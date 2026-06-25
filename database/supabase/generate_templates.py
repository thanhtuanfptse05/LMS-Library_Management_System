import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Define directories and template path
TEMPLATE_DIR = os.path.join("database", "supabase", "templates")
if not os.path.exists(TEMPLATE_DIR):
    os.makedirs(TEMPLATE_DIR)
BOOK_TEMPLATE_PATH = os.path.join(TEMPLATE_DIR, "template_import_books.xlsx")

# 25 Books data matching categories and tags in seed data
books_data = [
    {
        "isbn": "9780134685991",
        "title": "Effective Java (3rd Edition)",
        "author": "Joshua Bloch",
        "publisher": "Addison-Wesley",
        "publicationYear": 2018,
        "price": 850000.0,
        "categories": "Computer Science",
        "tags": "Textbook;Coding;Reference"
    },
    {
        "isbn": "9780132350884",
        "title": "Clean Code",
        "author": "Robert C. Martin",
        "publisher": "Prentice Hall",
        "publicationYear": 2008,
        "price": 650000.0,
        "categories": "Computer Science",
        "tags": "Textbook;Coding"
    },
    {
        "isbn": "9780201633610",
        "title": "Design Patterns: Elements of Reusable Object-Oriented Software",
        "author": "Erich Gamma, Richard Helm, Ralph Johnson, John Vlissides",
        "publisher": "Addison-Wesley",
        "publicationYear": 1994,
        "price": 950000.0,
        "categories": "Computer Science",
        "tags": "Reference;Theoretical"
    },
    {
        "isbn": "9781119560822",
        "title": "Corporate Finance (12th Edition)",
        "author": "Stephen Ross, Randolph Westerfield, Jeffrey Jaffe",
        "publisher": "Wiley",
        "publicationYear": 2019,
        "price": 1200000.0,
        "categories": "Economics & Finance",
        "tags": "Textbook;CaseStudy"
    },
    {
        "isbn": "9780321356680",
        "title": "Algorithms (4th Edition)",
        "author": "Robert Sedgewick, Kevin Wayne",
        "publisher": "Addison-Wesley",
        "publicationYear": 2011,
        "price": 900000.0,
        "categories": "Computer Science",
        "tags": "Textbook;Coding;Theoretical"
    },
    {
        "isbn": "9780073523323",
        "title": "Macroeconomics (9th Edition)",
        "author": "N. Gregory Mankiw",
        "publisher": "Worth Publishers",
        "publicationYear": 2015,
        "price": 800000.0,
        "categories": "Economics & Finance",
        "tags": "Textbook;Theoretical"
    },
    {
        "isbn": "9780199232741",
        "title": "Introduction to Politics",
        "author": "Robert Garner, Peter Ferdinand, Stephanie Lawson",
        "publisher": "Oxford University Press",
        "publicationYear": 2016,
        "price": 550000.0,
        "categories": "Politics & International Relations",
        "tags": "Introduction;Textbook"
    },
    {
        "isbn": "9781449331818",
        "title": "Designing Data-Intensive Applications",
        "author": "Martin Kleppmann",
        "publisher": "O'Reilly Media",
        "publicationYear": 2017,
        "price": 1100000.0,
        "categories": "Computer Science",
        "tags": "Reference;Advanced"
    },
    {
        "isbn": "9780132145374",
        "title": "Artificial Intelligence: A Modern Approach (3rd Edition)",
        "author": "Stuart Russell, Peter Norvig",
        "publisher": "Pearson",
        "publicationYear": 2010,
        "price": 1500000.0,
        "categories": "Computer Science",
        "tags": "Textbook;AI_Driven;Advanced"
    },
    {
        "isbn": "9780321573513",
        "title": "Introduction to Electrodynamics (4th Edition)",
        "author": "David J. Griffiths",
        "publisher": "Pearson",
        "publicationYear": 2012,
        "price": 750000.0,
        "categories": "Physics",
        "tags": "Textbook;Theoretical"
    },
    {
        "isbn": "9780134093413",
        "title": "Campbell Biology (11th Edition)",
        "author": "Lisa A. Urry, Michael L. Cain",
        "publisher": "Pearson",
        "publicationYear": 2016,
        "price": 1600000.0,
        "categories": "Biology & Ecology",
        "tags": "Textbook;Reference"
    },
    {
        "isbn": "9780470503201",
        "title": "Introduction to Physics (9th Edition)",
        "author": "John D. Cutnell, Kenneth W. Johnson",
        "publisher": "Wiley",
        "publicationYear": 2012,
        "price": 700000.0,
        "categories": "Physics",
        "tags": "Introduction;Textbook"
    },
    {
        "isbn": "9780078021558",
        "title": "Organic Chemistry (9th Edition)",
        "author": "Francis Carey, Robert Giuliano",
        "publisher": "McGraw-Hill Education",
        "publicationYear": 2013,
        "price": 1300000.0,
        "categories": "Chemistry",
        "tags": "Textbook;Experimental"
    },
    {
        "isbn": "9780136006633",
        "title": "Microbiology: An Introduction (10th Edition)",
        "author": "Gerard J. Tortora, Berdell R. Funke",
        "publisher": "Pearson",
        "publicationYear": 2009,
        "price": 1420000.0,
        "categories": "Biology & Ecology;Medicine & Health Sciences",
        "tags": "Textbook;LabManual"
    },
    {
        "isbn": "9780470646151",
        "title": "Principles of Anatomy and Physiology (13th Edition)",
        "author": "Gerard J. Tortora, Bryan H. Derrickson",
        "publisher": "Wiley",
        "publicationYear": 2011,
        "price": 1800000.0,
        "categories": "Medicine & Health Sciences",
        "tags": "Textbook;Reference"
    },
    {
        "isbn": "9781285057903",
        "title": "Introduction to Business (5th Edition)",
        "author": "Jeff Madura",
        "publisher": "Cengage Learning",
        "publicationYear": 2013,
        "price": 600000.0,
        "categories": "Business Administration",
        "tags": "Introduction;Textbook"
    },
    {
        "isbn": "9780133098754",
        "title": "Marketing Management (15th Edition)",
        "author": "Philip Kotler, Kevin Lane Keller",
        "publisher": "Pearson",
        "publicationYear": 2015,
        "price": 1150000.0,
        "categories": "Business Administration",
        "tags": "Textbook;CaseStudy;Strategic"
    },
    {
        "isbn": "9781133312789",
        "title": "General, Organic, and Biological Chemistry",
        "author": "H. Stephen Stoker",
        "publisher": "Cengage Learning",
        "publicationYear": 2015,
        "price": 980000.0,
        "categories": "Chemistry;Pharmacy & Biochemistry",
        "tags": "Textbook;LabManual"
    },
    {
        "isbn": "9780199571123",
        "title": "Introduction to Law",
        "author": "Jaap Hage, Bram Akkermans",
        "publisher": "Springer",
        "publicationYear": 2014,
        "price": 670000.0,
        "categories": "Law & Legal Studies",
        "tags": "Introduction;Textbook"
    },
    {
        "isbn": "9781305073036",
        "title": "Principles of Psychology",
        "author": "S. Marc Breedlove",
        "publisher": "Oxford University Press",
        "publicationYear": 2015,
        "price": 850000.0,
        "categories": "Psychology",
        "tags": "Textbook;Behavioral"
    },
    {
        "isbn": "9780321838964",
        "title": "Designing the User Interface (6th Edition)",
        "author": "Ben Shneiderman, Catherine Plaisant",
        "publisher": "Pearson",
        "publicationYear": 2016,
        "price": 930000.0,
        "categories": "Computer Science;Arts & Design",
        "tags": "Reference;Visual_Design"
    },
    {
        "isbn": "9780133914641",
        "title": "Software Engineering (10th Edition)",
        "author": "Ian Sommerville",
        "publisher": "Pearson",
        "publicationYear": 2015,
        "price": 950000.0,
        "categories": "Computer Science",
        "tags": "Textbook;Management"
    },
    {
        "isbn": "9781491950296",
        "title": "Kafka: The Definitive Guide",
        "author": "Neha Narkhede, Gwen Shapira, Todd Palino",
        "publisher": "O'Reilly Media",
        "publicationYear": 2017,
        "price": 650000.0,
        "categories": "Computer Science",
        "tags": "Reference;Advanced"
    },
    {
        "isbn": "9781501257285",
        "title": "Soft Skills: The software developer's life manual",
        "author": "John Sonmez",
        "publisher": "Manning Publications",
        "publicationYear": 2014,
        "price": 450000.0,
        "categories": "Soft Skills",
        "tags": "Reference;Management"
    },
    {
        "isbn": "9780073523859",
        "title": "Introduction to Algorithms (3rd Edition)",
        "author": "Thomas H. Cormen, Charles E. Leiserson, Ronald L. Rivest, Clifford Stein",
        "publisher": "MIT Press",
        "publicationYear": 2009,
        "price": 1400000.0,
        "categories": "Computer Science",
        "tags": "Textbook;Theoretical;Advanced"
    }
]

# Generate copies for each book (2 copies per book)
copies_data = []
for book in books_data:
    isbn = book["isbn"]
    for i in range(1, 3):  # 2 copies
        barcode = f"BC{isbn}-{i:02d}"
        location = f"Kệ CS-{isbn[-4:]}-{i}" if "Computer Science" in book["categories"] else f"Kệ GEN-{isbn[-4:]}-{i}"
        copies_data.append({
            "isbn": isbn,
            "barcode": barcode,
            "location": location
        })

# Create new workbook
wb = openpyxl.Workbook()

# Sheet 1: Books
ws_books = wb.active
ws_books.title = "Books"

# Column headers
book_headers = ["isbn", "title", "author", "publisher", "publicationYear", "price", "categories", "tags"]
ws_books.append(book_headers)

# Styles
font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
border_thin = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

# Apply header styles for Books
for col_idx, header in enumerate(book_headers, 1):
    cell = ws_books.cell(row=1, column=col_idx)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

# Fill Books data
for row_idx, book in enumerate(books_data, 2):
    ws_books.cell(row=row_idx, column=1, value=book["isbn"]).alignment = align_center
    ws_books.cell(row=row_idx, column=2, value=book["title"]).alignment = align_left
    ws_books.cell(row=row_idx, column=3, value=book["author"]).alignment = align_left
    ws_books.cell(row=row_idx, column=4, value=book["publisher"]).alignment = align_left
    ws_books.cell(row=row_idx, column=5, value=book["publicationYear"]).alignment = align_center
    
    price_cell = ws_books.cell(row=row_idx, column=6, value=book["price"])
    price_cell.number_format = '0.00'
    price_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    ws_books.cell(row=row_idx, column=7, value=book["categories"]).alignment = align_left
    ws_books.cell(row=row_idx, column=8, value=book["tags"]).alignment = align_left
    
    for col_idx in range(1, 9):
        ws_books.cell(row=row_idx, column=col_idx).border = border_thin

# Sheet 2: BookCopies
ws_copies = wb.create_sheet(title="BookCopies")
copy_headers = ["isbn", "barcode", "location"]
ws_copies.append(copy_headers)

# Apply header styles for BookCopies
for col_idx, header in enumerate(copy_headers, 1):
    cell = ws_copies.cell(row=1, column=col_idx)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

# Fill BookCopies data
for row_idx, copy in enumerate(copies_data, 2):
    ws_copies.cell(row=row_idx, column=1, value=copy["isbn"]).alignment = align_center
    ws_copies.cell(row=row_idx, column=2, value=copy["barcode"]).alignment = align_center
    ws_copies.cell(row=row_idx, column=3, value=copy["location"]).alignment = align_left
    
    for col_idx in range(1, 4):
        ws_copies.cell(row=row_idx, column=col_idx).border = border_thin

# Auto fit columns
for ws in [ws_books, ws_copies]:
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Save template
wb.save(BOOK_TEMPLATE_PATH)
print(f"Generated {BOOK_TEMPLATE_PATH} successfully with 25 books and {len(copies_data)} book copies.")
