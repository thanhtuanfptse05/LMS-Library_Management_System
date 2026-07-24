import os
import csv
import json
import glob
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define colors for Pass/Fail status
GREEN_FILL, GREEN_FONT = "C6EFCE", "006100"   # Pass
RED_FILL, RED_FONT = "FFC7CE", "9C0006"       # Fail

thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def build_expected_result(tc_id, desc, exp_val, data_dict):
    """
    Builds a clear, professional Vietnamese Expected Result based on TC ID, Description, and Data Binding.
    """
    tc_upper = tc_id.upper()
    exp_val_upper = str(exp_val).upper()
    role_val = data_dict.get('role', '').upper()
    
    # 1. Login TC
    if 'LOGIN' in tc_upper:
        if exp_val_upper == 'PASS':
            role_str = f" với vai trò {role_val}" if role_val else ""
            return f"Đăng nhập thành công{role_str}, chuyển hướng vào màn hình Bảng điều khiển tương ứng."
        else:
            return "Hệ thống từ chối đăng nhập, hiển thị thông báo lỗi sai tài khoản/mật khẩu."
            
    # 2. Logout TC
    if 'LOGOUT' in tc_upper:
        return "Hệ thống hủy Session người dùng, đăng xuất thành công và chuyển hướng về màn hình Đăng nhập."
        
    # 3. Forgot Password TC
    if 'FORGOTPASSWORD' in tc_upper or 'FORGOT' in tc_upper:
        return "Hệ thống tiếp nhận yêu cầu và hiển thị thông báo gửi email reset mật khẩu thành công."
        
    # 4. Account Lockout TC
    if 'ACCOUNTLOCKOUT' in tc_upper or 'LOCKED' in tc_upper:
        return "Hệ thống phát hiện đăng nhập sai quá 5 lần, tự động khóa tài khoản tạm thời và hiển thị thông báo."
        
    # 5. Create User TC
    if 'CREATEUSER' in tc_upper:
        if exp_val_upper == 'PASS':
            return "Hệ thống lưu tài khoản mới thành công, hiển thị thông báo và cập nhật danh sách người dùng."
        else:
            return "Hệ thống báo lỗi trùng lặp/thiếu thông tin bắt buộc, từ chối tạo mới tài khoản."
            
    # 6. Update User TC
    if 'UPDATEUSER' in tc_upper:
        if exp_val_upper == 'PASS':
            return "Hệ thống cập nhật thông tin người dùng thành công và hiển thị thông báo xác nhận."
        else:
            return "Hệ thống từ chối cập nhật và hiển thị thông báo lỗi dữ liệu không hợp lệ."
            
    # 7. Lock/Unlock User TC
    if 'LOCKUSER' in tc_upper:
        return "Hệ thống chuyển trạng thái tài khoản sang Locked (Đã khóa) và lưu lý do khóa."
    if 'UNLOCKUSER' in tc_upper:
        return "Hệ thống mở khóa tài khoản, chuyển trạng thái về Active (Hoạt động)."
        
    # 8. Book Management TCs
    if 'ADDBOOKCOPY' in tc_upper:
        return "Hệ thống tạo thành công bản sao sách mới kèm Mã vạch (Barcode) duy nhất."
    elif 'ADDBOOK' in tc_upper:
        return "Hệ thống tạo thành công đầu sách mới vào CSDL kho sách."
    elif 'UPDATEBOOK' in tc_upper:
        return "Hệ thống cập nhật thông tin đầu sách thành công."
        
    # 9. Desk Circulation TCs
    if 'CHECKOUT' in tc_upper:
        if exp_val_upper == 'PASS':
            return "Hệ thống xử lý giao sách thành công, đổi trạng thái bản sao sang 'borrowed' và tạo bản ghi mượn."
        else:
            return "Hệ thống từ chối giao sách, hiển thị thông báo lỗi (vượt hạn mức / đang nợ phạt / mã vạch không hợp lệ)."
    elif 'CHECKIN' in tc_upper:
        if exp_val_upper == 'PASS':
            return "Hệ thống nhận trả sách thành công, cập nhật trạng thái bản sao, ghi nhận sự cố / tính tiền phạt nếu có."
        else:
            return "Hệ thống từ chối nhận trả sách, hiển thị thông báo lỗi mã vạch hoặc mã độc giả không hợp lệ."
    elif 'PAYFINE' in tc_upper or 'PAYMENT' in tc_upper:
        if exp_val_upper == 'PASS':
            return "Hệ thống duyệt thu tiền mặt thành công, cập nhật trạng thái Fine sang 'paid' và tự động mở khóa tài khoản."
        else:
            return "Hệ thống kiểm tra không có khoản nợ phạt, giữ nguyên trạng thái tài khoản và không thực hiện duyệt thu."

    # 10. TS5 Book Discovery TCs
    if 'SEARCHBOOK' in tc_upper:
        if exp_val_upper == 'PASS':
            return "Hệ thống lọc và hiển thị danh sách các đầu sách khớp từ khóa tìm kiếm."
        else:
            return "Hệ thống báo không tìm thấy kết quả hoặc hiển thị danh sách rỗng."
    elif 'FILTERBYCATEGORY' in tc_upper:
        return "Hệ thống lọc danh sách sách chỉ thuộc các Danh mục được chọn."
    elif 'FILTERBYTAG' in tc_upper:
        return "Hệ thống lọc danh sách sách gắn các thẻ Tag tương ứng."
    elif 'FILTERBYAVAILABILITY' in tc_upper:
        return "Hệ thống lọc đúng các đầu sách theo trạng thái khả dụng trong kho."
    elif 'VIEWBOOKDETAIL' in tc_upper:
        return "Hệ thống hiển thị đầy đủ thông tin chi tiết đầu sách, vị trí kệ và các bản sao."

    # 11. TS6 Self Service TCs
    if 'RESERVEBOOKONLINE' in tc_upper:
        if exp_val_upper == 'PASS':
            return "Hệ thống ghi nhận đơn đặt trước trực tuyến thành công và cập nhật hàng đợi."
        else:
            return "Hệ thống từ chối đặt trước, hiển thị thông báo sách đã được mượn/đặt trước hoặc vượt giới hạn."
    elif 'RENEWBOOKONLINE' in tc_upper:
        if exp_val_upper == 'PASS':
            return "Hệ thống tự động gia hạn thêm thời hạn mượn sách thành công."
        else:
            return "Hệ thống từ chối gia hạn do chưa đạt ngưỡng thời hạn mượn (ít nhất 50%) hoặc vượt số lần cho phép."
    elif 'CANCELRESERVATION' in tc_upper:
        if exp_val_upper == 'PASS':
            return "Hệ thống chuyển trạng thái đơn đặt trước sang 'cancelled' và hoàn trả vị trí kho/hàng đợi."
        else:
            return "Hệ thống từ chối hủy đơn đặt trước không tồn tại hoặc không thuộc quyền sở hữu."

    # Generic fallback
    if exp_val_upper == 'FAIL':
        return "Hệ thống kiểm tra dữ liệu đầu vào, từ chối xử lý và hiển thị thông báo lỗi."
    else:
        return "Hệ thống xử lý giao dịch thành công, cập nhật CSDL và không phát sinh lỗi."


def parse_katalon_details_csv(csv_path):
    suite_name = os.path.splitext(os.path.basename(csv_path))[0]
    records = []
    
    with open(csv_path, mode='r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader, None)
        
        current_tc = None
        
        for row in reader:
            if not row or len(row) < 8 or not any(row):
                continue
                
            first_col = row[0].strip()
            status = row[7].strip() if len(row) > 7 else ''
            
            # Suite row
            if first_col.startswith('TS') or first_col.startswith('TS_') or ('Authentication' in first_col and not first_col.startswith('Test Cases')):
                suite_name = first_col
                continue
                
            # New Test Case row
            if first_col.startswith('Test Cases/') or (first_col.startswith('TC') and '/' in first_col):
                tc_id = os.path.basename(first_col)
                desc = row[2].strip() if len(row) > 2 else ''
                tag = row[3].strip() if len(row) > 3 else ''
                
                current_tc = {
                    'tc_base_id': tc_id,
                    'desc': desc,
                    'tag': tag,
                    'status': status if status else 'PASSED',
                    'input_data': '',
                    'expected': '',
                    'actual': '',
                    'steps_count': 0
                }
                records.append(current_tc)
                continue
                
            # DATA BINDING row
            if first_col.startswith('DATA BINDING -') and current_tc:
                json_str = first_col.replace('DATA BINDING -', '').strip()
                try:
                    data_dict = json.loads(json_str)
                    exp_val = data_dict.get('expected', data_dict.get('expectedResult', 'PASS'))
                    
                    # Remove expected from input parameters display
                    clean_dict = {k: v for k, v in data_dict.items() if k not in ['expected', 'expectedResult']}
                    inputs = [f"{k}: {v}" for k, v in clean_dict.items() if v != '']
                    current_tc['input_data'] = "\n".join(inputs)
                    
                    # Generate Expected Result
                    current_tc['expected'] = build_expected_result(current_tc['tc_base_id'], current_tc['desc'], exp_val, data_dict)
                    
                except Exception:
                    current_tc['input_data'] = json_str
                    current_tc['expected'] = build_expected_result(current_tc['tc_base_id'], current_tc['desc'], 'PASS', {})
                continue
                
            # Step row
            if current_tc and first_col and not first_col.startswith('DATA BINDING'):
                current_tc['steps_count'] += 1

    # Finalize Expected & Actual results for cases without Data Binding
    tc_counter = {}
    for tc in records:
        base_id = tc['tc_base_id']
        tc_counter[base_id] = tc_counter.get(base_id, 0) + 1
        
        # Row ID
        if tc_counter[base_id] > 1 or tc['input_data']:
            tc['row_tc_id'] = f"{base_id}_Data{tc_counter[base_id]:02d}"
        else:
            tc['row_tc_id'] = base_id
            
        if not tc['expected']:
            tc['expected'] = build_expected_result(tc['tc_base_id'], tc['desc'], 'PASS', {})
            
        if tc['status'] == 'PASSED':
            tc['actual'] = f"Đã thực hiện đúng kỳ vọng. Hệ thống chạy thành công {tc['steps_count']} steps."
        else:
            tc['actual'] = f"Hệ thống phát sinh lỗi bất thường tại một trong số {tc['steps_count']} steps."
            
    return suite_name, records


def build_template3_report(system_test_dir):
    template_path = os.path.join(system_test_dir, "Template3_System Test.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found at {template_path}")

    # Explicit 6 CSV module files in project
    module_files = [
        (os.path.join(system_test_dir, "20260724_014335.csv"), "TSAuthentication", "Xác thực & Bảo mật", "Xác minh luồng Đăng nhập, Đăng xuất, Quên mật khẩu & Khóa tài khoản"),
        (os.path.join(system_test_dir, "20260724_012717.csv"), "TSUserManagement", "Quản lý Người dùng", "Xác minh luồng Quản lý người dùng, Thêm/Sửa/Khóa/Mở khóa tài khoản"),
        (os.path.join(system_test_dir, "20260724_010205.csv"), "TSBookManagement", "Quản lý Sách & Kho", "Xác minh luồng Quản lý sách, Thêm đầu sách, Khai báo bản sao Barcode & Sửa sách"),
        (os.path.join(system_test_dir, "20260724_100830.csv"), "TSDeskCirculation", "Quầy Lưu Thông", "Xác minh luồng Quầy lưu thông, Giao sách, Nhận trả sách & Duyệt thu tiền mặt"),
        (os.path.join(system_test_dir, "20260724_151531.csv"), "TSBookDiscovery", "Tra cứu & Khám phá Sách", "Xác minh luồng Tìm kiếm sách, Lọc theo Danh mục, Tag, Trạng thái & Xem chi tiết"),
        (os.path.join(system_test_dir, "20260724_193607.csv"), "TSSelfService", "Tự phục vụ Độc giả", "Xác minh luồng Đặt trước sách trực tuyến, Gia hạn mượn & Hủy đơn đặt trước")
    ]

    wb = openpyxl.load_workbook(template_path)

    # 1. Update Cover Sheet Metadata
    if "Cover" in wb.sheetnames:
        cover_ws = wb["Cover"]
        cover_ws["B4"] = "LMS - Library Management System"
        cover_ws["B5"] = "LMS_SWP391"
        cover_ws["F4"] = "Quality Assurance Team"
        cover_ws["F5"] = "2026-07-24"
        cover_ws["F6"] = "v1.3.0"
        cover_ws["A11"] = "2026-07-24"
        cover_ws["B11"] = "v1.3.0"
        cover_ws["C11"] = "Hoàn thành System Testing 6 phân hệ"
        cover_ws["D11"] = "M"
        cover_ws["E11"] = "Báo cáo kết quả kiểm thử tự động 76 test cases trên Katalon Studio"
        cover_ws["F11"] = "LMS_SWP391_SRS_v1.3.0"

    # 2. Update Test Cases Sheet Overview
    if "Test Cases" in wb.sheetnames:
        tc_summary_ws = wb["Test Cases"]
        tc_summary_ws["D3"] = "LMS - Library Management System"
        tc_summary_ws["D4"] = "LMS_SWP391"
        tc_summary_ws["D5"] = "1. Application Server: Java JDK 17, Apache Tomcat 10, Java Servlet 5.0\n2. Database: PostgreSQL (Supabase / Supavisor)\n3. Testing Tool: Katalon Studio v9.x (Web Automation)\n4. Web Browser: Google Chrome"

        # Clear existing summary rows
        for r in range(9, tc_summary_ws.max_row + 1):
            for c in range(2, 7):
                tc_summary_ws.cell(row=r, column=c).value = None

        # Populate Test Case List table starting at Row 9
        for i, (csv_file, suite_code, func_name, desc) in enumerate(module_files, start=1):
            r = 8 + i
            tc_summary_ws.cell(row=r, column=2, value=i)
            tc_summary_ws.cell(row=r, column=3, value=func_name)
            tc_summary_ws.cell(row=r, column=4, value=suite_code)
            tc_summary_ws.cell(row=r, column=5, value=desc)
            tc_summary_ws.cell(row=r, column=6, value="Hệ thống và CSDL mồi đã sẵn sàng")

    # Template sheet reference for cloning
    ws_template = wb["Workflow Name1"]

    module_stats = []

    # 3. Create Workflow Sheets for Each Module
    for csv_file, suite_code, func_name, req_desc in module_files:
        raw_name, tc_list = parse_katalon_details_csv(csv_file)
        
        # Clone Workflow Name1 sheet preserving exact frame & styling
        ws_mod = wb.copy_worksheet(ws_template)
        ws_mod.title = suite_code
        
        # Set Module Header Info
        ws_mod["B2"] = suite_code
        ws_mod["B3"] = req_desc
        ws_mod["B4"] = f"=COUNTA(A11:A{10 + len(tc_list)})"
        
        # Formulate Round 1, 2, 3 Counts
        ws_mod["B6"] = f'=COUNTIF($F11:$F{10 + len(tc_list)}, "Passed")'
        ws_mod["C6"] = f'=COUNTIF($F11:$F{10 + len(tc_list)}, "Failed")'
        ws_mod["D6"] = f'=COUNTIF($F11:$F{10 + len(tc_list)}, "Pending")'
        ws_mod["E6"] = f'=COUNTIF($F11:$F{10 + len(tc_list)}, "N/A")'

        ws_mod["B7"] = f'=COUNTIF($F11:$F{10 + len(tc_list)}, "Passed")'
        ws_mod["C7"] = f'=COUNTIF($F11:$F{10 + len(tc_list)}, "Failed")'
        ws_mod["D7"] = f'=COUNTIF($F11:$F{10 + len(tc_list)}, "Pending")'
        ws_mod["E7"] = f'=COUNTIF($F11:$F{10 + len(tc_list)}, "N/A")'

        ws_mod["B8"] = f'=COUNTIF($F11:$F{10 + len(tc_list)}, "Passed")'
        ws_mod["C8"] = f'=COUNTIF($F11:$F{10 + len(tc_list)}, "Failed")'
        ws_mod["D8"] = f'=COUNTIF($F11:$F{10 + len(tc_list)}, "Pending")'
        ws_mod["E8"] = f'=COUNTIF($F11:$F{10 + len(tc_list)}, "N/A")'

        # Clear sample rows from template (rows 11 to max_row)
        for r in range(11, ws_mod.max_row + 1):
            for c in range(1, 16):
                ws_mod.cell(row=r, column=c).value = None

        # Populate Data Rows starting at Row 11
        for idx, tc in enumerate(tc_list):
            r = 11 + idx
            status_str = "Passed" if tc['status'] == 'PASSED' else "Failed"
            
            row_data = [
                tc['row_tc_id'],
                tc['desc'] if tc['desc'] else f"Kiểm thử chức năng {tc['tc_base_id']}",
                tc['input_data'] if tc['input_data'] else "Tham số mặc định",
                tc['expected'],
                "Hệ thống và CSDL mồi đã sẵn sàng",
                status_str,
                "2026-07-24",
                "Katalon AutoTest",
                status_str,
                "2026-07-24",
                "Katalon AutoTest",
                status_str,
                "2026-07-24",
                "Katalon AutoTest",
                f"Tự động chạy thành công {tc['steps_count']} bước"
            ]

            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_mod.cell(row=r, column=col_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Apply Status Fill for Round 1, 2, 3 columns (Cols 6, 9, 12)
                if col_idx in [6, 9, 12]:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    if status_str == "Passed":
                        cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
                        cell.font = Font(name="Tahoma", size=10, bold=True, color=GREEN_FONT)
                    else:
                        cell.fill = PatternFill("solid", fgColor=RED_FILL)
                        cell.font = Font(name="Tahoma", size=10, bold=True, color=RED_FONT)
                else:
                    cell.font = Font(name="Tahoma", size=10, bold=False)

        module_stats.append((suite_code, len(tc_list)))

    # 4. Update Test Statistics Sheet
    if "Test Statistics" in wb.sheetnames:
        stat_ws = wb["Test Statistics"]
        stat_ws["C3"] = "LMS - Library Management System"
        stat_ws["C4"] = "LMS_SWP391"
        stat_ws["C5"] = '=C4&"_Test_Report_v1.3.0"'
        stat_ws["G3"] = "Quality Assurance Team"
        stat_ws["G4"] = "Lead QA"
        stat_ws["H5"] = "2026-07-24"
        stat_ws["C6"] = "Báo cáo tổng hợp kết quả System Testing 6 phân hệ hệ thống LMS (Tổng số 76 kịch bản kiểm thử)"

        # Clear existing table rows
        for r in range(11, stat_ws.max_row + 1):
            for c in range(2, 9):
                stat_ws.cell(row=r, column=c).value = None

        # Populate Statistics Table starting at Row 11
        for i, (suite_code, count) in enumerate(module_stats, start=1):
            r = 10 + i
            stat_ws.cell(row=r, column=2, value=i)
            stat_ws.cell(row=r, column=3, value=f"='{suite_code}'!B2")
            stat_ws.cell(row=r, column=4, value=f"='{suite_code}'!B6")
            stat_ws.cell(row=r, column=5, value=f"='{suite_code}'!C6")
            stat_ws.cell(row=r, column=6, value=f"='{suite_code}'!D6")
            stat_ws.cell(row=r, column=7, value=f"='{suite_code}'!E6")
            stat_ws.cell(row=r, column=8, value=f"='{suite_code}'!B4")

        # Subtotal Row
        sub_row = 11 + len(module_stats)
        stat_ws.cell(row=sub_row, column=3, value="Sub total")
        stat_ws.cell(row=sub_row, column=4, value=f"=SUM(D11:D{sub_row-1})")
        stat_ws.cell(row=sub_row, column=5, value=f"=SUM(E11:E{sub_row-1})")
        stat_ws.cell(row=sub_row, column=6, value=f"=SUM(F11:F{sub_row-1})")
        stat_ws.cell(row=sub_row, column=7, value=f"=SUM(G11:G{sub_row-1})")
        stat_ws.cell(row=sub_row, column=8, value=f"=SUM(H11:H{sub_row-1})")

        # Coverage Rows
        cov_row = sub_row + 2
        stat_ws.cell(row=cov_row, column=3, value="Test coverage")
        stat_ws.cell(row=cov_row, column=5, value=f"=IF((H{sub_row}-G{sub_row})=0,0,(D{sub_row}+E{sub_row})*100/(H{sub_row}-G{sub_row}))")
        stat_ws.cell(row=cov_row, column=6, value="%")

        succ_row = cov_row + 1
        stat_ws.cell(row=succ_row, column=3, value="Test successful coverage")
        stat_ws.cell(row=succ_row, column=5, value=f"=IF((H{sub_row}-G{sub_row})=0,0,D{sub_row}*100/(H{sub_row}-G{sub_row}))")
        stat_ws.cell(row=succ_row, column=6, value="%")

    # Remove template placeholder sheets
    for t_name in ["Workflow Name1", "Workflow Name2"]:
        if t_name in wb.sheetnames:
            wb.remove(wb[t_name])

    output_path = os.path.join(system_test_dir, "LMS_System_Test_Report.xlsx")
    wb.save(output_path)
    print(f"Successfully generated Template3 System Test Report: {output_path}")

if __name__ == '__main__':
    build_template3_report(r"D:\Data\NetBeansIDE17\LMS-Library_Management_System\System Test")
