import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_DIR = r"d:\Data\NetBeansIDE17\LMS-Library_Management_System\System Test\Data System Test"

def create_excel_file(filename, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    for row_idx, row_data in enumerate(rows, start=2):
        ws.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = align_left if col_idx == len(row_data) else align_center
            cell.border = border
            cell.font = Font(name="Segoe UI", size=10)
            
    filepath = os.path.join(DATA_DIR, filename)
    wb.save(filepath)
    print(f"Created {filepath} successfully!")

def main():
    # 1. DataSearchBook.xlsx (TCSearchBook)
    headers_search = ["keyword", "expected", "Ghi chú"]
    rows_search = [
        ["Java", "PASS", "1. Tìm kiếm sách theo từ khóa tên sách 'Java'"],
        ["Kiểm Thử", "PASS", "2. Tìm kiếm sách theo từ khóa 'Kiểm Thử'"],
        ["978-0134685991", "PASS", "3. Tìm kiếm sách chính xác theo mã ISBN '978-0134685991'"],
        ["Joshua Bloch", "PASS", "4. Tìm kiếm sách theo tên tác giả 'Joshua Bloch'"],
        ["KhongTonTai123456", "FAIL", "5. Tìm kiếm từ khóa không tồn tại trên hệ thống (báo rỗng/không tìm thấy)"],
        ["", "PASS", "6. Để trống từ khóa bấm Tìm kiếm (hệ thống trả về toàn bộ danh mục sách)"]
    ]
    create_excel_file("DataSearchBook.xlsx", headers_search, rows_search)

    # 2. DataFilterByCategory.xlsx (TCFilterByCategory)
    headers_cat = ["categoryId", "expected", "Ghi chú"]
    rows_cat = [
        ["1", "PASS", "1. Lọc danh sách sách theo Danh mục ID = 1"],
        ["9999", "FAIL", "2. Lọc theo Danh mục ID không tồn tại trên hệ thống"]
    ]
    create_excel_file("DataFilterByCategory.xlsx", headers_cat, rows_cat)

    # 3. DataFilterByTag.xlsx (TCFilterByTag)
    headers_tag = ["tagId", "expected", "Ghi chú"]
    rows_tag = [
        ["1", "PASS", "1. Lọc danh sách sách theo Thẻ Tag ID = 1"],
        ["9999", "FAIL", "2. Lọc theo Thẻ Tag ID không tồn tại trên hệ thống"]
    ]
    create_excel_file("DataFilterByTag.xlsx", headers_tag, rows_tag)

    # 4. DataFilterByAvailability.xlsx (TCFilterByAvailability)
    headers_avail = ["availableOnly", "expected", "Ghi chú"]
    rows_avail = [
        ["true", "PASS", "1. Lọc sách chỉ lấy các đầu sách sẵn có trong kho (availableOnly=true)"],
        ["false", "PASS", "2. Lọc tất cả đầu sách kể cả hết bản sao mượn (availableOnly=false)"]
    ]
    create_excel_file("DataFilterByAvailability.xlsx", headers_avail, rows_avail)

    # 5. DataViewBookDetail.xlsx (TCViewBookDetail)
    headers_detail = ["bookId", "expected", "Ghi chú"]
    rows_detail = [
        ["991", "PASS", "1. Xem chi tiết đầu sách hợp lệ 'Giáo Trình Kiểm Thử LMS 2026' (bookId=991)"],
        ["9999", "FAIL", "2. Xem chi tiết đầu sách không tồn tại (bookId=9999)"]
    ]
    create_excel_file("DataViewBookDetail.xlsx", headers_detail, rows_detail)

if __name__ == "__main__":
    main()
