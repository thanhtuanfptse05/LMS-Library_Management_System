import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_DIR = r"d:\Data\NetBeansIDE17\LMS-Library_Management_System\System Test\Data System Test"

def create_excel_file(filename, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    for row_idx, row_data in enumerate(rows, start=2):
        ws.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = align_left if col_idx == len(row_data) else align_center
            cell.border = border
            cell.font = Font(name="Segoe UI", size=10)
            
    filepath = os.path.join(DATA_DIR, filename)
    wb.save(filepath)
    print(f"Created {filepath} successfully!")

def main():
    # 1. DataReserveBookOnline.xlsx (TCReserveBookOnline)
    headers_reserve = ["bookId", "expected", "Ghi chú"]
    rows_reserve = [
        ["991", "PASS", "1. Đặt trước sách thành công khi còn bản sao / có thể xếp hàng chờ"],
        ["9999", "FAIL", "2. Đặt trước đầu sách không tồn tại"]
    ]
    create_excel_file("DataReserveBookOnline.xlsx", headers_reserve, rows_reserve)

    # 2. DataRenewBookOnline.xlsx (TCRenewBookOnline)
    headers_renew = ["borrowRecordId", "expected", "Ghi chú"]
    rows_renew = [
        ["9902", "PASS", "1. Độc giả gửi yêu cầu gia hạn hợp lệ cho sách đang mượn"],
        ["9999", "FAIL", "2. Gia hạn bản ghi mượn không tồn tại / quá số lần cho phép"]
    ]
    create_excel_file("DataRenewBookOnline.xlsx", headers_renew, rows_renew)

    # 3. DataCancelReservation.xlsx (TCCancelReservation)
    headers_cancel = ["reservationId", "expected", "Ghi chú"]
    rows_cancel = [
        ["9901", "PASS", "1. Hủy đơn đặt trước sách thành công (chuyển cancelled)"],
        ["9999", "FAIL", "2. Hủy đơn đặt trước không tồn tại"]
    ]
    create_excel_file("DataCancelReservation.xlsx", headers_cancel, rows_cancel)

if __name__ == "__main__":
    main()
