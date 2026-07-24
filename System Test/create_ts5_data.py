import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_data_book_discovery():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DataBookDiscovery"
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    headers = [
        "TC_ID", "keyword", "categoryId", "tagId", "availableOnly", "expected", "Description"
    ]
    
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    data = [
        ("TCSearchBook_Data01", "Java", "", "", "false", "PASS", "Tìm kiếm sách theo từ khóa tên sách 'Java'"),
        ("TCSearchBook_Data02", "Kiểm Thử", "", "", "false", "PASS", "Tìm kiếm sách theo từ khóa 'Kiểm Thử'"),
        ("TCSearchBook_Data03", "978-0134685991", "", "", "false", "PASS", "Tìm kiếm sách chính xác theo mã ISBN '978-0134685991'"),
        ("TCSearchBook_Data04", "Joshua Bloch", "", "", "false", "PASS", "Tìm kiếm sách theo tên tác giả 'Joshua Bloch'"),
        ("TCSearchBook_Data05", "KhongTonTai123456", "", "", "false", "EMPTY", "Tìm kiếm từ khóa không tồn tại trên hệ thống"),
        ("TCFilterByCategory_Data01", "", "1", "", "false", "PASS", "Lọc danh sách sách theo Danh mục ID = 1"),
        ("TCFilterByTag_Data01", "", "", "1", "false", "PASS", "Lọc danh sách sách theo Thẻ Tag ID = 1"),
        ("TCFilterByAvailability_Data01", "", "", "", "true", "PASS", "Lọc sách chỉ lấy các đầu sách sẵn có trong kho (availableOnly=true)"),
        ("TCViewBookDetail_Data01", "Giáo Trình Kiểm Thử LMS 2026", "", "", "false", "PASS", "Click xem chi tiết đầu sách 'Giáo Trình Kiểm Thử LMS 2026'")
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        ws.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = align_center if col_idx in [1, 3, 4, 5, 6] else align_left
            cell.border = border
            cell.font = Font(name="Segoe UI", size=10)
            
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 50
    
    output_path = r"d:\Data\NetBeansIDE17\LMS-Library_Management_System\System Test\DataBookDiscovery.xlsx"
    wb.save(output_path)
    print(f"Created {output_path} successfully!")

if __name__ == "__main__":
    create_data_book_discovery()
